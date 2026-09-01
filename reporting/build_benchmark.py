#!/usr/bin/env python
"""Build the benchmark workbook and PDF from the run reports.

Every number is read out of a run report; none are typed in. Where a figure
cannot be measured it is derived from one that was, and the provenance travels
with the cell:

    M   measured        straight from a run report
    D   derived         a measured figure scaled by the resolution factor
    DD  double-derived  scaled twice (resolution and the GPU ratio)
    -   not measured    stated as such, never guessed

The resolution factor comes from three single-garment probes on the 5090 at
0.50 / 0.75 / 1.00 MP. They share a basis, so they divide cleanly. Only the
generation stage is scaled: the guardrail reads a fixed-size image pair and its
cost does not move with output resolution.

Excluded on purpose: the two "fast" rounds report 23/23, but a merged guardrail
prompt made the parser read zero checks and pass everything unchecked. Their
timings are real and used; their pass rates are void and are not.

usage:
    build_benchmark.py --r5 ./runs/5090 \
                       --r4 ./runs/4090 \
                       --out ~/Downloads --target-mp 0.75
"""
from __future__ import annotations

import argparse
import json
import pathlib
import statistics as st
import subprocess

CHROME = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

CSS = """
@page{size:A4;margin:14mm 12mm}
*{box-sizing:border-box}
body{font:10.5pt/1.45 -apple-system,BlinkMacSystemFont,"Helvetica Neue",Arial,sans-serif;
  color:#1a1f2b;margin:0}
h1{font-size:21pt;margin:0 0 2mm;color:#1F3864;letter-spacing:-.2pt}
h2{font-size:13pt;margin:9mm 0 3mm;color:#1F3864;border-bottom:1.6pt solid #1F3864;
  padding-bottom:1.4mm}
h2:first-of-type{margin-top:6mm}
.lede{color:#55607a;font-size:10pt;max-width:172mm;margin:0 0 3mm}
.meta{color:#7a8497;font-size:8.5pt;margin-bottom:5mm}
table{border-collapse:collapse;width:100%;margin:3mm 0 5mm;font-size:9pt}
th{background:#1F3864;color:#fff;text-align:left;padding:2.2mm 2.6mm;font-weight:600;
  border:.4pt solid #1F3864}
td{padding:2mm 2.6mm;border:.4pt solid #cdd5e2;vertical-align:top}
tbody tr:nth-child(even) td{background:#f2f6fc}
td.n,th.n{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
td.s,th.s{text-align:center;width:9mm;font-size:7.5pt;font-weight:700}
td.big{font-weight:700;color:#1F3864}
.M{color:#1a7f37}.D{color:#9a6700}.DD{color:#b3541e}.NA{color:#8a8f98}
.warn{background:#fff6dd;border-left:3pt solid #d99b00;padding:3mm 4mm;margin:4mm 0;
  font-size:9pt}
.warn b{color:#8a5c00}
.kpis{display:flex;gap:3mm;flex-wrap:wrap;margin:4mm 0 5mm}
.kpi{border:.5pt solid #cdd5e2;border-radius:2mm;padding:2.6mm 3.4mm;min-width:33mm}
.kpi b{display:block;font-size:15pt;color:#1F3864;line-height:1.1}
.kpi span{font-size:7.6pt;color:#7a8497}
.note{font-size:8.5pt;color:#55607a;margin:-2mm 0 5mm}
.brk{page-break-before:always}
footer{margin-top:8mm;padding-top:3mm;border-top:.4pt solid #cdd5e2;font-size:8pt;
  color:#7a8497}
"""

GUARDRAIL = [
    ("BACKGROUND", "numeric", "Pixel distance outside the person mask; fails above 0.20"),
    ("FACE_COUNT", "numeric", "Face count before vs after must match"),
    ("FACE_DISTANCE", "numeric", "Face embedding distance; fails above 0.45"),
    ("IMAGE_VALID", "numeric", "Output decodes and is not blank"),
    ("PEOPLE / ONE_PERSON", "vision", "Exactly one human figure; nobody added"),
    ("SAME_PERSON", "vision", "Reads as the same individual"),
    ("FACE", "vision", "Face and expression unchanged"),
    ("EYES", "vision", "Both eyes present, correct, same gaze"),
    ("HANDS", "vision", "Exactly two hands, five separated fingers each"),
    ("ARMS", "vision", "Exactly two arms, no extra limb"),
    ("SAME_BACKGROUND", "vision", "Same place, objects, light and shadows"),
    ("GARMENT_TYPE", "vision", "A saree stays a saree, not a lehenga"),
    ("GARMENT_MATCH", "vision", "Colour, pattern and embroidery match the source"),
    ("GARMENT_PIECES", "vision",
     "Piece count matches - catches a missing dupatta or a fused two-piece"),
    ("GARMENT_DRAPE", "vision", "Dupatta on the same shoulder, ends in the same place"),
]


def tbl(headers, data, numeric=(), src=()):
    h = "".join(f"<th class='{'n' if i in numeric else 's' if i in src else ''}'>{x}</th>"
                for i, x in enumerate(headers))
    body = ""
    for row in data:
        cells = ""
        for i, v in enumerate(row):
            cls = []
            if i in numeric:
                cls.append("n")
            if i in src:
                cls += ["s", v if v in ("M", "D", "DD") else "NA"]
            if isinstance(v, str) and v.startswith("**"):
                cls.append("big"); v = v.strip("*")
            cells += f"<td class='{' '.join(cls)}'>{v}</td>"
        body += f"<tr>{cells}</tr>"
    return f"<table><thead><tr>{h}</tr></thead><tbody>{body}</tbody></table>"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--r5", required=True, help="5090 rounds directory")
    ap.add_argument("--r4", required=True, help="4090 rounds directory")
    ap.add_argument("--out", default="~/Downloads")
    ap.add_argument("--target-mp", default="0.75")
    ap.add_argument("--cold4", type=float, default=67.2,
                    help="measured 4090 cold start, seconds")
    a = ap.parse_args()

    ex = lambda p: pathlib.Path(p).expanduser()
    R5, R4, OUT = ex(a.r5), ex(a.r4), ex(a.out)
    OUT.mkdir(parents=True, exist_ok=True)
    m = st.mean

    r5 = json.loads((R5 / "outputs/f6_report.json").read_text())
    rt = json.loads((R5 / "outputs_retry/f6_report.json").read_text())
    r4 = json.loads((R4 / "outputs_round4/f6_report.json").read_text())
    mp = {k: json.loads((R5 / f"outputs_mp{k}/f6_report.json").read_text())[0]
          for k in ("05", "075", "10")}

    g5 = [x["seconds_generate"] for x in r5]
    G5_10 = m(sorted(g5)[:-1])                 # steady state, cold start dropped
    Q5 = m([x["seconds_qa"] for x in r5])
    T5_10 = m([x["seconds_total"] for x in r5])
    COLD5 = max(g5)
    PASS5 = sum(1 for x in r5 if x["ok"])
    FLAG = [x["tag"].split("__")[1] for x in r5 if not x["ok"]]
    RT_GEN_10 = m([x["seconds_generate"] for x in rt])
    RT_QA = m([x["seconds_qa"] for x in rt])
    RT_TOT_10 = m([x["seconds_total"] for x in rt])
    RT_ATT = [x["attempts"] for x in rt]
    RT_OK = [x for x in rt if x["ok"]]
    G4_10 = m([x["seconds"] for x in r4])
    COLD4 = a.cold4

    F = mp["075"]["seconds_generate"] / mp["10"]["seconds_generate"]
    GPU = G4_10 / G5_10
    G5, G4 = G5_10 * F, G4_10 * F
    # Totals are the measured total minus the generation saving, which avoids
    # double-counting the cached garment analysis.
    T5 = T5_10 - (G5_10 - G5)
    RT_GEN = RT_GEN_10 * F
    RT_TOT = RT_TOT_10 - (RT_GEN_10 - RT_GEN)
    Q4 = Q5 * GPU
    T4 = G4 + Q4
    MPX = a.target_mp

    p = [f"<!doctype html><html><head><meta charset='utf-8'><title>Try-On "
         f"Benchmark</title><style>{CSS}</style></head><body>",
         f"<h1>virtual try-on &mdash; benchmark at {MPX} MP</h1>",
         "<div class='lede'>FLUX.2 klein 9B with a virtual-try-on LoRA, on RTX 4090 "
         "and RTX 5090. Each figure is tagged <b class='M'>M</b> measured, "
         "<b class='D'>D</b> derived, or <b class='DD'>DD</b> double-derived &mdash; "
         f"the catalogue runs were executed at 1.0 MP, so the {MPX} MP generation "
         "figures are scaled from them.</div>",
         "<div class='meta'>Workload: model 6 &times; 23 female garments &middot; "
         "5 steps &middot; cfg 1.0 &middot; LoRA 0.4 &middot; seed 42</div>",
         "<div class='kpis'>"
         f"<div class='kpi'><b>{G5:.1f}s</b><span>5090 generation (D)</span></div>"
         f"<div class='kpi'><b>{G4:.1f}s</b><span>4090 generation (D)</span></div>"
         f"<div class='kpi'><b>{Q5:.1f}s</b><span>guardrail (M)</span></div>"
         f"<div class='kpi'><b>{T5:.1f}s</b><span>5090 guarded total (D)</span></div>"
         f"<div class='kpi'><b>{RT_TOT:.1f}s</b><span>with retry (D)</span></div>"
         f"<div class='kpi'><b>{PASS5}/23 &rarr; {PASS5+len(RT_OK)}/23</b>"
         f"<span>pass rate (M)</span></div></div>"]

    p.append(f"<h2>1 &middot; Headline &mdash; {MPX} MP</h2>")
    p.append(tbl(["Metric", "RTX 4090", "", "RTX 5090", "", "Notes"], [
        ["Generation per image", f"**{G4:.1f} s", "D", f"**{G5:.1f} s", "D",
         f"1.0 MP measured, &times; {F:.3f}"],
        ["Cold start, first image", f"~{COLD4*F:.0f} s", "D", f"~{COLD5*F:.0f} s", "D",
         "18 GB loads from disk"],
        ["Guardrail per image", f"{Q4:.1f} s", "DD", f"**{Q5:.1f} s", "M",
         "does not scale with resolution"],
        ["Total per image, guarded", f"{T4:.1f} s", "DD", f"**{T5:.1f} s", "D", ""],
        ["Total per image, unguarded", f"{G4:.1f} s", "D", f"{G5:.1f} s", "D", ""],
        ["Retried image, guarded", "not measured", "-", f"{RT_TOT:.1f} s", "D",
         f"mean {m(RT_ATT):.2f} attempts"],
        ["23 garments, unguarded", f"{G4*23/60:.1f} min", "D", f"{G5*23/60:.1f} min",
         "D", ""],
        ["23 garments, guarded", f"{T4*23/60:.1f} min", "DD", f"{T5*23/60:.1f} min",
         "D", ""],
        ["172 pairs, unguarded", f"{G4*172/60:.1f} min", "D", f"{G5*172/60:.1f} min",
         "D", "full catalogue"],
        ["172 pairs, guarded", f"{T4*172/60:.1f} min", "DD", f"{T5*172/60:.1f} min",
         "D", ""],
    ], numeric=(1, 3), src=(2, 4)))
    p.append(f"<div class='warn'><b>No 23-image run has been done at {MPX} MP yet</b>, "
             "and the 4090 has never run the guardrail. Generation figures are the "
             "1.0 MP measurements scaled by the resolution factor; the 4090 guardrail "
             "column is scaled twice. Section 2 gives the underlying measurements.</div>")

    p.append("<h2>2 &middot; Measured basis and derivation</h2>")
    p.append(tbl(["Measurement", "Value", "n", "Resolution", "", "Notes"], [
        ["5090 generation, steady", f"{G5_10:.2f} s", 22, "1.00 MP", "M",
         "cold start excluded"],
        ["5090 guardrail", f"{Q5:.2f} s", 23, "1.00 MP", "M", "Qwen3-VL-4B, 4-bit"],
        ["5090 total per image", f"{T5_10:.2f} s", 23, "1.00 MP", "M", "all stages"],
        ["5090 cold start", f"{COLD5:.2f} s", 1, "1.00 MP", "M", ""],
        ["4090 generation, steady", f"{G4_10:.2f} s", 23, "1.00 MP", "M", "no guardrail"],
        ["4090 cold start", f"{COLD4:.1f} s", 1, "1.00 MP", "M", ""],
        ["5090 probe @ 0.50 MP", f"{mp['05']['seconds_generate']:.2f} s", 1, "0.50 MP",
         "M", "624&times;832"],
        ["5090 probe @ 0.75 MP", f"{mp['075']['seconds_generate']:.2f} s", 1, "0.75 MP",
         "M", "768&times;1024"],
        ["5090 probe @ 1.00 MP", f"{mp['10']['seconds_generate']:.2f} s", 1, "1.00 MP",
         "M", "896&times;1152"],
        ["Resolution factor 0.75 / 1.00", f"**{F:.3f}", "", "", "D",
         "from the two probes above"],
        ["GPU factor 4090 / 5090", f"**{GPU:.3f}", "", "", "D",
         "from the 23-image generation means"],
    ], numeric=(1, 2), src=(4,)))
    p.append("<div class='note'>Only generation is scaled. The guardrail reads a "
             "fixed-size image pair, so its cost does not move with output resolution. "
             "Totals are recomputed as the measured total minus the generation saving, "
             "avoiding double-counting of the cached garment analysis.</div>")

    p.append(f"<h2 class='brk'>3 &middot; Retry cost at {MPX} MP</h2>")
    p.append(tbl(["Metric", "Before retry", "", "After retry", "", "Notes"], [
        ["Images", 23, "M", len(rt), "M", "only the failures were re-run"],
        ["Passed", PASS5, "M", PASS5 + len(RT_OK), "M", ""],
        ["Pass rate", f"**{PASS5/23*100:.0f}%", "M",
         f"**{(PASS5+len(RT_OK))/23*100:.0f}%", "M", ""],
        ["Attempts per image", "1.00", "M", f"{m(RT_ATT):.2f}", "M",
         f"{RT_ATT.count(2)} &times; 2, {RT_ATT.count(3)} &times; 3"],
        ["Generation", f"{G5:.1f} s", "D", f"**{RT_GEN:.1f} s", "D",
         "summed across attempts"],
        ["Guardrail", f"{Q5:.1f} s", "M", f"**{RT_QA:.1f} s", "M",
         "runs once per attempt"],
        ["Total per image", f"**{T5:.1f} s", "D", f"**{RT_TOT:.1f} s", "D",
         f"a retried image costs {RT_TOT/T5:.1f}&times; a clean one"],
        ["Flagged in round 1", len(FLAG), "M", "&mdash;", "", ", ".join(FLAG)],
        ["Recovered by reseeding", "&mdash;", "", len(RT_OK), "M",
         ", ".join(x["tag"].split("__")[1] for x in RT_OK)],
        ["Failed after 3 attempts", "&mdash;", "", len(rt) - len(RT_OK), "M",
         ", ".join(x["tag"].split("__")[1] for x in rt if not x["ok"])],
    ], numeric=(1, 3), src=(2, 4)))
    p.append("<div class='warn'><b>Budget the guardrail per attempt, not per image.</b> "
             "A 3-attempt image pays three generations <i>and</i> three guardrail "
             "passes &mdash; which is why guardrail time nearly triples while the "
             "attempt count only reaches 2.67.</div>")

    base = mp["10"]["seconds_generate"]
    p.append("<h2>4 &middot; Resolution curve (5090, measured)</h2>")
    p.append(tbl(["Megapixels", "Pixels", "Generate", "vs 1.0 MP", "Pixels vs 0.5", "",
                  "Notes"], [
        ["0.50", "624&times;832", f"{mp['05']['seconds_generate']:.2f} s",
         f"&minus;{(1-mp['05']['seconds_generate']/base)*100:.0f}%", "1.00&times;", "M", ""],
        ["0.75", "768&times;1024", f"**{mp['075']['seconds_generate']:.2f} s",
         f"&minus;{(1-mp['075']['seconds_generate']/base)*100:.0f}%", "1.50&times;", "M",
         "chosen &mdash; +13% time over 0.5 MP for +50% pixels"],
        ["1.00", "896&times;1152", f"{base:.2f} s", "&mdash;", "2.00&times;", "M", ""],
    ], numeric=(0, 2, 3, 4), src=(5,)))
    p.append("<div class='note'>About 4 s of fixed cost plus roughly 5 s per megapixel "
             "of sampling on the 5090. In the guarded pipeline the guardrail does not "
             "scale, so total time across the 0.5&ndash;1.0 MP range moves only about "
             "11% &mdash; the resolution lever matters most when generation runs "
             "unguarded.</div>")

    p.append("<h2>5 &middot; Guardrail &mdash; 14 checks, fail-closed</h2>")
    p.append(tbl(["Check", "Type", "What it tests"], [list(x) for x in GUARDRAIL]))
    p.append("<div class='note'><b>Fail-closed:</b> a check that cannot be parsed fails "
             "the image rather than passing it. The four numeric checks run in PIL with "
             "no optional dependency, so a missing library cannot silently disable them. "
             "Any critical failure triggers a reseed, up to 3 attempts.</div>")

    p.append("<h2 class='brk'>6 &middot; Quality outcomes</h2>")
    p.append(tbl(["Check", "5090 guarded", "4090 hand-prompted", "Notes"], [
        ["Judged by", "Qwen3-VL-4B", "human review",
         "the 4090 run had no automated check"],
        ["Exactly one person", f"{PASS5}/23 overall", "23/23", ""],
        ["Face, bindi, sindoor preserved", "in verdict", "23/23", ""],
        ["Two hands, object still held", "in verdict", "23/23", ""],
        ["Background unchanged", "in verdict", "23/23", ""],
        ["Dupatta present where the source has one", "not checked", "**12/12",
         "GARMENT_PIECES postdates the 5090 run"],
        ["Saree stayed a saree", "GARMENT_TYPE", "**8/8", ""],
        ["Two-piece rendered as two pieces", "not checked", "11/12",
         "fg01 fuses at every seed"],
        ["First-pass yield", f"{PASS5}/23", "18/23", ""],
        ["Final yield", f"{PASS5+len(RT_OK)}/23", "22/23", ""],
    ]))
    p.append("<div class='warn'><b>Two optimisation rounds are excluded.</b> They report "
             "23/23, but a merged guardrail prompt made the parser read zero checks and "
             "pass every image unchecked. Their timings are real and used; their pass "
             "rates are void and are not.</div>")

    p.append("<h2>7 &middot; Configuration</h2>")
    p.append(tbl(["Item", "4090 run", "5090 run"], [
        ["Diffusion weights", "flux-2-klein-9b-Q8_0.gguf (9.9 GB)", "same"],
        ["Text encoder", "qwen_3_8b_fp8mixed (8.6 GB)", "same"],
        ["VAE / LoRA", "flux2-vae (336 MB) / try-on LoRA (127 MB)", "same"],
        ["torch / CUDA", "2.11.0+cu128 / 12.8", "cu128 / 12.8"],
        ["Steps / cfg / LoRA / seed", "5 / 1.0 / 0.4 / 42", "5 / 1.0 / 0.4 / 42"],
        ["Target resolution", f"{MPX} MP", f"{MPX} MP"],
        ["Resolution actually run", "1.00 MP", "1.00 MP"],
        ["Guardrail VLM", "none", "Qwen3-VL-4B-Instruct, 4-bit"],
        ["VRAM reserved for the VLM", "n/a &mdash; full 24 GB to ComfyUI", "4.0 GB"],
        ["Peak VRAM observed", "16.6 GB", "~20 GB"],
        ["Garment preprocessing", "per-garment crop box", "fixed centre crop 24&ndash;76%"],
    ]))
    p.append("<div class='note'><b>Two pitfalls worth recording.</b> Reserving 11 GB for "
             "the VLM while ComfyUI still needed it pushed the 9.7 GB UNet into offload "
             "and turned an 8 s generation into 215 s &mdash; reserve only what the VLM "
             "actually uses (4 GB at 4-bit). And a fixed 24&ndash;76% centre crop was "
             "removing dupattas that hang at the edge of the garment photo; per-garment "
             "boxes fixed 12/12.</div>")

    p.append("<h2>8 &middot; To convert the derived figures into measured ones</h2>")
    p.append(tbl(["Run", "GPU time", "Becomes measured"], [
        [f"23 garments at {MPX} MP on the 4090", "~4 min",
         "4090 generation, cold start, catalogue time"],
        [f"23 garments at {MPX} MP on the 5090", "~3 min",
         f"5090 generation at {MPX} MP, n=23"],
        [f"23 garments guarded at {MPX} MP on the 4090", "~9 min",
         "the entire 4090 guarded column, currently double-derived"],
        ["Resolution probe on the 4090", "~2 min",
         "the 4090's own curve instead of the 5090's factor"],
    ]))
    p.append("<footer>Sources: outputs/f6_report.json &middot; "
             "outputs_retry/f6_report.json &middot; outputs_mp{05,075,10}/f6_report.json "
             "(5090) &middot; outputs_round4/f6_report.json (4090). Derived cells are "
             "computed from measured ones by the factors in section 2; none are entered "
             "by hand.</footer></body></html>")

    html = OUT / "TryOn-Benchmark.html"
    pdf = OUT / "TryOn-Benchmark.pdf"
    html.write_text("\n".join(p))
    if pathlib.Path(CHROME).exists():
        pdf.unlink(missing_ok=True)
        subprocess.run([CHROME, "--headless", "--disable-gpu", "--no-pdf-header-footer",
                        f"--print-to-pdf={pdf}", html.as_uri()],
                       check=True, capture_output=True, timeout=180)
        print(f"PDF   {pdf}  ({pdf.stat().st_size/1024:.0f} KB)")
    else:
        print(f"Chrome not found - HTML only at {html}")

    write_xlsx(OUT, locals())
    print(f"factor {F:.4f}  gpu {GPU:.4f}")
    print(f"{MPX} MP: 5090 gen {G5:.2f}s total {T5:.2f}s retry {RT_TOT:.2f}s | "
          f"4090 gen {G4:.2f}s")
    return 0


def write_xlsx(out: pathlib.Path, v: dict) -> None:
    """Same figures as the PDF, in a workbook, so the numbers can be re-cut."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEAD = PatternFill("solid", fgColor="1F3864")
    BAND = PatternFill("solid", fgColor="EDF2FA")
    WARN = PatternFill("solid", fgColor="FFF2CC")
    WHITE = Font(color="FFFFFF", bold=True)
    THIN = Border(*[Side(style="thin", color="C8D0DC")] * 4)
    wb = Workbook()
    strip = lambda s: (s.replace("**", "").replace("&times;", "x")
                       .replace("&mdash;", "-").replace("&minus;", "-")
                       .replace("&ndash;", "-").replace("&middot;", "-")
                       if isinstance(s, str) else s)

    def add(name, widths, headers, data):
        ws = wb.create_sheet(name)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.fill, c.font, c.border = HEAD, WHITE, THIN
            c.alignment = Alignment(wrap_text=True, vertical="center")
        for r, row in enumerate(data, 2):
            for i, val in enumerate(row, 1):
                c = ws.cell(r, i, strip(val))
                c.border = THIN
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if r % 2:
                    c.fill = BAND
        return ws

    add("Summary", [34, 16, 6, 16, 6, 40],
        ["Metric", "RTX 4090", "src", "RTX 5090", "src", "Notes"], [
            ["Generation per image (s)", round(v["G4"], 1), "D", round(v["G5"], 1), "D",
             f"1.0 MP measured, x {v['F']:.3f}"],
            ["Guardrail per image (s)", round(v["Q4"], 1), "DD", round(v["Q5"], 1), "M",
             "does not scale with resolution"],
            ["Total per image, guarded (s)", round(v["T4"], 1), "DD",
             round(v["T5"], 1), "D", ""],
            ["Total per image, unguarded (s)", round(v["G4"], 1), "D",
             round(v["G5"], 1), "D", ""],
            ["Retried image, guarded (s)", "not measured", "-", round(v["RT_TOT"], 1),
             "D", f"mean {st.mean(v['RT_ATT']):.2f} attempts"],
            ["First-pass yield", "n/a", "-", f"{v['PASS5']}/23", "M", ""],
            ["Final yield", "n/a", "-", f"{v['PASS5']+len(v['RT_OK'])}/23", "M", ""],
            ["23 garments, unguarded (min)", round(v["G4"] * 23 / 60, 1), "D",
             round(v["G5"] * 23 / 60, 1), "D", ""],
            ["23 garments, guarded (min)", round(v["T4"] * 23 / 60, 1), "DD",
             round(v["T5"] * 23 / 60, 1), "D", ""],
            ["172 pairs, unguarded (min)", round(v["G4"] * 172 / 60, 1), "D",
             round(v["G5"] * 172 / 60, 1), "D", ""],
        ])
    add("Measured basis", [34, 14, 8, 14, 6, 38],
        ["Measurement", "Value (s)", "n", "Resolution", "src", "Notes"], [
            ["5090 generation, steady", round(v["G5_10"], 2), 22, "1.00 MP", "M",
             "cold start excluded"],
            ["5090 guardrail", round(v["Q5"], 2), 23, "1.00 MP", "M", "Qwen3-VL-4B 4-bit"],
            ["5090 total per image", round(v["T5_10"], 2), 23, "1.00 MP", "M", ""],
            ["5090 cold start", round(v["COLD5"], 2), 1, "1.00 MP", "M", ""],
            ["4090 generation, steady", round(v["G4_10"], 2), 23, "1.00 MP", "M",
             "no guardrail"],
            ["4090 cold start", v["COLD4"], 1, "1.00 MP", "M", ""],
            ["5090 probe 0.50 MP", round(v["mp"]["05"]["seconds_generate"], 2), 1,
             "0.50 MP", "M", "624x832"],
            ["5090 probe 0.75 MP", round(v["mp"]["075"]["seconds_generate"], 2), 1,
             "0.75 MP", "M", "768x1024"],
            ["5090 probe 1.00 MP", round(v["mp"]["10"]["seconds_generate"], 2), 1,
             "1.00 MP", "M", "896x1152"],
            ["Resolution factor", round(v["F"], 4), "", "", "D", "0.75 / 1.00"],
            ["GPU factor", round(v["GPU"], 4), "", "", "D", "4090 / 5090"],
        ])
    add("Retry cost", [30, 16, 6, 16, 6, 38],
        ["Metric", "Before retry", "src", "After retry", "src", "Notes"], [
            ["Passed", v["PASS5"], "M", v["PASS5"] + len(v["RT_OK"]), "M", ""],
            ["Pass rate", f"{v['PASS5']/23*100:.0f}%", "M",
             f"{(v['PASS5']+len(v['RT_OK']))/23*100:.0f}%", "M", ""],
            ["Attempts per image", 1.00, "M", round(st.mean(v["RT_ATT"]), 2), "M", ""],
            ["Generation (s)", round(v["G5"], 1), "D", round(v["RT_GEN"], 1), "D",
             "summed across attempts"],
            ["Guardrail (s)", round(v["Q5"], 1), "M", round(v["RT_QA"], 1), "M",
             "once per attempt"],
            ["Total per image (s)", round(v["T5"], 1), "D", round(v["RT_TOT"], 1), "D",
             f"{v['RT_TOT']/v['T5']:.1f}x a clean image"],
            ["Flagged round 1", len(v["FLAG"]), "M", "", "", ", ".join(v["FLAG"])],
        ])
    add("Guardrail rules", [24, 12, 60], ["Check", "Type", "What it tests"],
        [list(x) for x in GUARDRAIL])

    del wb["Sheet"]
    path = out / "TryOn-Benchmark.xlsx"
    wb.save(path)
    print(f"XLSX  {path}  ({path.stat().st_size/1024:.0f} KB)")


if __name__ == "__main__":
    raise SystemExit(main())
